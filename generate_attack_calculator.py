#!/usr/bin/env python3
"""
Generate Excel workbook for attack calculation in the Security Planspiel.
Creates one worksheet per attack where measures can be entered and results are calculated.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# Load configuration
with open('simulation_config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

# Extract data
measures = config['measures']
waves = config['waves']
attacks = config['attacks']
base_cia = config['base_cia']

# Styling
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
RESULT_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
BONUS_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
ATTACK_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
CIA_C_FILL = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
CIA_I_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
CIA_A_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_attack_worksheet(wb, wave_data, attack_data):
    """Create a worksheet for a specific attack/wave."""
    wave_id = wave_data['id']
    wave_name = wave_data['name']
    attack_id = wave_data['attack_id']

    ws = wb.create_sheet(title=f"Welle {wave_id} - {wave_name}")

    # Title
    ws['A1'] = f"Angriff Welle {wave_id}: {wave_name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:L1')

    # Attack info
    ws['A3'] = "Angriffsparameter"
    ws['A3'].font = TITLE_FONT

    ws['A4'] = "Basis-Schweregrad:"
    ws['B4'] = attack_data['base_severity']
    ws['C4'] = f"Schadenseinheit: {attack_data['s_unit']}k EUR"
    ws['A5'] = "Mitigation Cap:"
    ws['B5'] = attack_data['mitigation_cap']
    ws['C5'] = f"Recovery erlaubt: {'Ja' if attack_data['allow_recovery'] else 'Nein'}"

    ws['A6'] = "CIA-Gewichtung:"
    ws['B6'] = f"C={wave_data['weights']['c']}, I={wave_data['weights']['i']}, A={wave_data['weights']['a']}"
    ws['A7'] = "E-Wert Ziel:"
    ws['B7'] = wave_data['e_threshold']
    ws['C7'] = f"KZ Bonus/Malus: +{wave_data['kz_bonus']} / {wave_data['kz_malus']}"

    for row in range(4, 8):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = ATTACK_FILL

    # Bonus measures info
    ws['A9'] = "Bonus-Maßnahmen für diesen Angriff:"
    ws['A9'].font = TITLE_FONT

    bonus_row = 10
    for bm in wave_data['bonus_measures']:
        ws[f'A{bonus_row}'] = f"{bm['measure']} (min. Level {bm['min_level']}): +{bm['bonus']} Reduktion"
        ws[f'B{bonus_row}'] = bm['description']
        ws[f'A{bonus_row}'].fill = BONUS_FILL
        bonus_row += 1

    # Measures input section
    start_row = bonus_row + 2
    ws[f'A{start_row}'] = "Maßnahmen-Eingabe"
    ws[f'A{start_row}'].font = TITLE_FONT

    # Headers
    header_row = start_row + 1
    headers = ['Maßnahme', 'Name', 'Level', 'C', 'I', 'A', 'Init (k€)', 'OPEX (k€)', 'Recovery']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if header == 'C':
            cell.fill = CIA_C_FILL
        elif header == 'I':
            cell.fill = CIA_I_FILL
        elif header == 'A':
            cell.fill = CIA_A_FILL

    # Add data validation for level input (0-3)
    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=3,
                        errorTitle="Ungültiger Level",
                        error="Bitte einen Wert zwischen 0 und 3 eingeben.")
    ws.add_data_validation(dv)

    # Measure rows
    measure_start_row = header_row + 1
    measure_ids = ['M1', 'M2', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8', 'M9', 'M10']

    for idx, m_id in enumerate(measure_ids):
        row = measure_start_row + idx
        m_data = measures[m_id]

        # Measure ID
        ws.cell(row=row, column=1, value=m_id).border = THIN_BORDER
        # Measure Name
        ws.cell(row=row, column=2, value=m_data['name']).border = THIN_BORDER
        # Level input cell (editable)
        level_cell = ws.cell(row=row, column=3, value=0)
        level_cell.fill = INPUT_FILL
        level_cell.border = THIN_BORDER
        level_cell.alignment = Alignment(horizontal='center')
        dv.add(level_cell)

        # CIA values (formulas based on level)
        level_ref = f'$C${row}'

        # Build CHOOSE formulas for CIA values
        c_vals = [m_data['levels'][str(l)]['cia']['c'] for l in range(4)]
        i_vals = [m_data['levels'][str(l)]['cia']['i'] for l in range(4)]
        a_vals = [m_data['levels'][str(l)]['cia']['a'] for l in range(4)]
        init_vals = [m_data['levels'][str(l)]['init'] for l in range(4)]
        opex_vals = [m_data['levels'][str(l)]['opex'] for l in range(4)]

        # Recovery values (only M4 has recovery)
        recovery_vals = []
        for l in range(4):
            lvl_data = m_data['levels'][str(l)]
            recovery_vals.append(lvl_data.get('recovery', 0))

        # C value
        c_cell = ws.cell(row=row, column=4)
        c_cell.value = f'=CHOOSE({level_ref}+1,{c_vals[0]},{c_vals[1]},{c_vals[2]},{c_vals[3]})'
        c_cell.border = THIN_BORDER
        c_cell.fill = CIA_C_FILL
        c_cell.alignment = Alignment(horizontal='center')

        # I value
        i_cell = ws.cell(row=row, column=5)
        i_cell.value = f'=CHOOSE({level_ref}+1,{i_vals[0]},{i_vals[1]},{i_vals[2]},{i_vals[3]})'
        i_cell.border = THIN_BORDER
        i_cell.fill = CIA_I_FILL
        i_cell.alignment = Alignment(horizontal='center')

        # A value
        a_cell = ws.cell(row=row, column=6)
        a_cell.value = f'=CHOOSE({level_ref}+1,{a_vals[0]},{a_vals[1]},{a_vals[2]},{a_vals[3]})'
        a_cell.border = THIN_BORDER
        a_cell.fill = CIA_A_FILL
        a_cell.alignment = Alignment(horizontal='center')

        # Init cost
        init_cell = ws.cell(row=row, column=7)
        init_cell.value = f'=CHOOSE({level_ref}+1,{init_vals[0]},{init_vals[1]},{init_vals[2]},{init_vals[3]})'
        init_cell.border = THIN_BORDER
        init_cell.alignment = Alignment(horizontal='center')

        # OPEX cost
        opex_cell = ws.cell(row=row, column=8)
        opex_cell.value = f'=CHOOSE({level_ref}+1,{opex_vals[0]},{opex_vals[1]},{opex_vals[2]},{opex_vals[3]})'
        opex_cell.border = THIN_BORDER
        opex_cell.alignment = Alignment(horizontal='center')

        # Recovery factor
        rec_cell = ws.cell(row=row, column=9)
        rec_cell.value = f'=CHOOSE({level_ref}+1,{recovery_vals[0]},{recovery_vals[1]},{recovery_vals[2]},{recovery_vals[3]})'
        rec_cell.border = THIN_BORDER
        rec_cell.alignment = Alignment(horizontal='center')

    # Totals row
    total_row = measure_start_row + len(measure_ids)
    ws.cell(row=total_row, column=1, value="SUMME").font = HEADER_FONT
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    ws.cell(row=total_row, column=2).border = THIN_BORDER
    ws.cell(row=total_row, column=3).border = THIN_BORDER

    # Sum formulas
    for col in [4, 5, 6, 7, 8]:
        cell = ws.cell(row=total_row, column=col)
        col_letter = get_column_letter(col)
        cell.value = f'=SUM({col_letter}{measure_start_row}:{col_letter}{total_row-1})'
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if col == 4:
            cell.fill = CIA_C_FILL
        elif col == 5:
            cell.fill = CIA_I_FILL
        elif col == 6:
            cell.fill = CIA_A_FILL

    # Max recovery formula
    max_rec_cell = ws.cell(row=total_row, column=9)
    max_rec_cell.value = f'=MAX(I{measure_start_row}:I{total_row-1})'
    max_rec_cell.font = HEADER_FONT
    max_rec_cell.border = THIN_BORDER
    max_rec_cell.alignment = Alignment(horizontal='center')

    # Calculation section
    calc_row = total_row + 3
    ws[f'A{calc_row}'] = "Berechnungsergebnis"
    ws[f'A{calc_row}'].font = TITLE_FONT

    # Team CIA (base + measures)
    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Basis-CIA:").font = HEADER_FONT
    ws.cell(row=calc_row, column=2, value=f"C={base_cia['c']}, I={base_cia['i']}, A={base_cia['a']}")

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Team CIA (Basis + Maßnahmen):").font = HEADER_FONT

    # Team C
    ws.cell(row=calc_row, column=2, value="C:")
    team_c_cell = ws.cell(row=calc_row, column=3)
    team_c_cell.value = f'={base_cia["c"]}+D{total_row}'
    team_c_cell.fill = RESULT_FILL
    team_c_cell.border = THIN_BORDER

    # Team I
    ws.cell(row=calc_row, column=4, value="I:")
    team_i_cell = ws.cell(row=calc_row, column=5)
    team_i_cell.value = f'={base_cia["i"]}+E{total_row}'
    team_i_cell.fill = RESULT_FILL
    team_i_cell.border = THIN_BORDER

    # Team A
    ws.cell(row=calc_row, column=6, value="A:")
    team_a_cell = ws.cell(row=calc_row, column=7)
    team_a_cell.value = f'={base_cia["a"]}+F{total_row}'
    team_a_cell.fill = RESULT_FILL
    team_a_cell.border = THIN_BORDER

    team_cia_row = calc_row

    # E-value calculation
    calc_row += 2
    ws.cell(row=calc_row, column=1, value="E-Wert Berechnung:").font = HEADER_FONT
    ws.cell(row=calc_row, column=2, value=f"({wave_data['weights']['c']}×C) + ({wave_data['weights']['i']}×I) + ({wave_data['weights']['a']}×A)")

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="E-Wert:").font = HEADER_FONT
    e_value_cell = ws.cell(row=calc_row, column=2)
    w = wave_data['weights']
    e_value_cell.value = f'=({w["c"]}*C{team_cia_row})+({w["i"]}*E{team_cia_row})+({w["a"]}*G{team_cia_row})'
    e_value_cell.fill = RESULT_FILL
    e_value_cell.border = THIN_BORDER
    e_value_cell.number_format = '0.00'
    e_value_row = calc_row

    ws.cell(row=calc_row, column=3, value=f"Ziel: {wave_data['e_threshold']}")
    ws.cell(row=calc_row, column=4, value="Erreicht:")
    target_cell = ws.cell(row=calc_row, column=5)
    target_cell.value = f'=IF(B{e_value_row}>={wave_data["e_threshold"]},"JA","NEIN")'
    target_cell.fill = RESULT_FILL
    target_cell.border = THIN_BORDER

    # Bonus reduction calculation
    calc_row += 2
    ws.cell(row=calc_row, column=1, value="Bonus-Reduktion (spezifische Maßnahmen):").font = HEADER_FONT

    bonus_formula_parts = []
    for bm in wave_data['bonus_measures']:
        m_num = int(bm['measure'][1:])  # Extract number from M1, M2, etc.
        level_row = measure_start_row + m_num - 1
        bonus_formula_parts.append(f'IF(C{level_row}>={bm["min_level"]},{bm["bonus"]},0)')

    calc_row += 1
    bonus_formula = '+'.join(bonus_formula_parts) if bonus_formula_parts else '0'
    ws.cell(row=calc_row, column=1, value="Bonus-Reduktion:").font = HEADER_FONT
    bonus_cell = ws.cell(row=calc_row, column=2)
    bonus_cell.value = f'={bonus_formula}'
    bonus_cell.fill = BONUS_FILL
    bonus_cell.border = THIN_BORDER
    bonus_row = calc_row

    # E-value based reduction
    calc_row += 2
    ws.cell(row=calc_row, column=1, value="E-Wert Reduktion:").font = HEADER_FONT
    ws.cell(row=calc_row, column=2, value=f"=MAX(0,(B{e_value_row}-{wave_data['e_threshold']})/{wave_data['e_divisor']})")
    e_reduction_row = calc_row

    # Total reduction (capped)
    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Gesamt-Reduktion:").font = HEADER_FONT
    total_red_cell = ws.cell(row=calc_row, column=2)
    total_red_cell.value = f'=B{e_reduction_row}+B{bonus_row}'
    total_red_cell.border = THIN_BORDER
    total_reduction_row = calc_row

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Reduktion (mit Cap):").font = HEADER_FONT
    capped_cell = ws.cell(row=calc_row, column=2)
    capped_cell.value = f'=MIN(B{total_reduction_row},{attack_data["mitigation_cap"]})'
    capped_cell.fill = RESULT_FILL
    capped_cell.border = THIN_BORDER
    capped_row = calc_row

    # Final severity
    calc_row += 2
    ws.cell(row=calc_row, column=1, value="Finaler Schweregrad:").font = HEADER_FONT
    severity_cell = ws.cell(row=calc_row, column=2)
    severity_cell.value = f'=MAX(0,{attack_data["base_severity"]}-B{capped_row})'
    severity_cell.fill = RESULT_FILL
    severity_cell.border = THIN_BORDER
    severity_row = calc_row

    # Damage before recovery
    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Schaden (vor Recovery):").font = HEADER_FONT
    damage_cell = ws.cell(row=calc_row, column=2)
    damage_cell.value = f'=B{severity_row}*{attack_data["s_unit"]}'
    damage_cell.border = THIN_BORDER
    damage_cell.number_format = '0 "k EUR"'
    damage_before_row = calc_row

    # Recovery factor (only if allowed)
    calc_row += 1
    if attack_data['allow_recovery']:
        ws.cell(row=calc_row, column=1, value="Recovery-Faktor:").font = HEADER_FONT
        recovery_factor_cell = ws.cell(row=calc_row, column=2)
        recovery_factor_cell.value = f'=I{total_row}'
        recovery_factor_cell.border = THIN_BORDER
        recovery_factor_cell.number_format = '0%'
        recovery_factor_row = calc_row

        calc_row += 1
        ws.cell(row=calc_row, column=1, value="Schaden (nach Recovery):").font = Font(bold=True, size=12, color='C00000')
        final_damage_cell = ws.cell(row=calc_row, column=2)
        final_damage_cell.value = f'=B{damage_before_row}*(1-B{recovery_factor_row})'
        final_damage_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        final_damage_cell.border = THIN_BORDER
        final_damage_cell.font = Font(bold=True, size=12, color='C00000')
        final_damage_cell.number_format = '0 "k EUR"'
    else:
        ws.cell(row=calc_row, column=1, value="Recovery nicht möglich").font = Font(italic=True)
        calc_row += 1
        ws.cell(row=calc_row, column=1, value="Finaler Schaden:").font = Font(bold=True, size=12, color='C00000')
        final_damage_cell = ws.cell(row=calc_row, column=2)
        final_damage_cell.value = f'=B{damage_before_row}'
        final_damage_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        final_damage_cell.border = THIN_BORDER
        final_damage_cell.font = Font(bold=True, size=12, color='C00000')
        final_damage_cell.number_format = '0 "k EUR"'

    # KZ impact
    calc_row += 2
    ws.cell(row=calc_row, column=1, value="Kundenvertrauen (KZ) Impact:").font = TITLE_FONT

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="KZ-Verlust durch Angriff:").font = HEADER_FONT
    kz_loss_cell = ws.cell(row=calc_row, column=2)
    kz_loss_cell.value = f'=-B{severity_row}*{attack_data["kz_unit"]}'
    kz_loss_cell.border = THIN_BORDER
    kz_loss_row = calc_row

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="KZ E-Wert Bonus/Malus:").font = HEADER_FONT
    kz_bonus_cell = ws.cell(row=calc_row, column=2)
    kz_bonus_cell.value = f'=IF(B{e_value_row}>={wave_data["e_threshold"]},{wave_data["kz_bonus"]},{wave_data["kz_malus"]})'
    kz_bonus_cell.border = THIN_BORDER
    kz_bonus_row = calc_row

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Gesamt KZ-Delta:").font = Font(bold=True, size=12)
    total_kz_cell = ws.cell(row=calc_row, column=2)
    total_kz_cell.value = f'=B{kz_loss_row}+B{kz_bonus_row}'
    total_kz_cell.fill = RESULT_FILL
    total_kz_cell.border = THIN_BORDER
    total_kz_cell.font = Font(bold=True, size=12)

    # Cost summary
    calc_row += 2
    ws.cell(row=calc_row, column=1, value="Kosten:").font = TITLE_FONT

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="Initiale Investition:").font = HEADER_FONT
    init_cost_cell = ws.cell(row=calc_row, column=2)
    init_cost_cell.value = f'=G{total_row}'
    init_cost_cell.border = THIN_BORDER
    init_cost_cell.number_format = '0 "k EUR"'

    calc_row += 1
    ws.cell(row=calc_row, column=1, value="OPEX pro Welle:").font = HEADER_FONT
    opex_cost_cell = ws.cell(row=calc_row, column=2)
    opex_cost_cell.value = f'=H{total_row}'
    opex_cost_cell.border = THIN_BORDER
    opex_cost_cell.number_format = '0 "k EUR"'

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    return ws


def create_overview_sheet(wb):
    """Create an overview sheet with general information."""
    ws = wb.active
    ws.title = "Übersicht"

    ws['A1'] = "Security Planspiel - Angriffs-Berechnungstabelle"
    ws['A1'].font = Font(bold=True, size=18)
    ws.merge_cells('A1:F1')

    ws['A3'] = "Anleitung"
    ws['A3'].font = TITLE_FONT

    instructions = [
        "1. Wählen Sie ein Arbeitsblatt für den jeweiligen Angriff (Welle 1, 2 oder 3)",
        "2. Geben Sie für jede Maßnahme (M1-M10) das gewählte Level ein (0-3)",
        "3. Die CIA-Werte, Kosten und Berechnungen werden automatisch aktualisiert",
        "4. Das Ergebnis zeigt den finalen Schweregrad und Schaden nach Mitigation",
        "",
        "Die gelb markierten Zellen sind Eingabefelder für die Maßnahmen-Level.",
        "Die grün markierten Zellen zeigen berechnete Ergebnisse."
    ]

    for i, text in enumerate(instructions):
        ws[f'A{4+i}'] = text

    # Wave summary
    ws['A13'] = "Angriffs-Übersicht"
    ws['A13'].font = TITLE_FONT

    headers = ['Welle', 'Name', 'Basis-Schweregrad', 'Schadenseinheit', 'E-Ziel', 'Max Schaden (ohne Mitigation)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for i, wave in enumerate(waves):
        row = 15 + i
        attack = attacks[wave['attack_id']]
        max_damage = attack['base_severity'] * attack['s_unit']

        ws.cell(row=row, column=1, value=wave['id']).border = THIN_BORDER
        ws.cell(row=row, column=2, value=wave['name']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=attack['base_severity']).border = THIN_BORDER
        ws.cell(row=row, column=4, value=f"{attack['s_unit']}k EUR").border = THIN_BORDER
        ws.cell(row=row, column=5, value=wave['e_threshold']).border = THIN_BORDER
        ws.cell(row=row, column=6, value=f"{max_damage}k EUR").border = THIN_BORDER

    # Measures overview
    ws['A20'] = "Maßnahmen-Übersicht (CIA-Werte pro Level)"
    ws['A20'].font = TITLE_FONT

    m_headers = ['ID', 'Name', 'L0 C/I/A', 'L1 C/I/A', 'L2 C/I/A', 'L3 C/I/A']
    for col, header in enumerate(m_headers, 1):
        cell = ws.cell(row=21, column=col, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for i, m_id in enumerate(['M1', 'M2', 'M3', 'M4', 'M5', 'M6', 'M7', 'M8', 'M9', 'M10']):
        row = 22 + i
        m_data = measures[m_id]

        ws.cell(row=row, column=1, value=m_id).border = THIN_BORDER
        ws.cell(row=row, column=2, value=m_data['name']).border = THIN_BORDER

        for lvl in range(4):
            cia = m_data['levels'][str(lvl)]['cia']
            cell = ws.cell(row=row, column=3+lvl, value=f"{cia['c']}/{cia['i']}/{cia['a']}")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25

    return ws


def main():
    """Generate the Excel workbook."""
    wb = Workbook()

    # Create overview sheet
    create_overview_sheet(wb)

    # Create worksheet for each wave/attack
    for wave in waves:
        attack_id = wave['attack_id']
        attack_data = attacks[attack_id]
        create_attack_worksheet(wb, wave, attack_data)

    # Save workbook
    output_file = 'angriff_berechnung.xlsx'
    wb.save(output_file)
    print(f"Excel-Datei erstellt: {output_file}")
    print(f"Arbeitsblätter: {[ws.title for ws in wb.worksheets]}")


if __name__ == '__main__':
    main()
